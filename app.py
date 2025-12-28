import streamlit as st
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime, timedelta
import plotly.express as px
import json
import traceback
from io import BytesIO
import os

# Excel/formatting
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None

# --- UI / Page config ---
st.set_page_config(page_title="CBHI Performance Tracker", layout="wide", page_icon="📈")

# --- 1. CONFIGURATION & PLAN DATA ---
PLANS = {
    "01 Merged Health Post": {"high": 453, "medium": 551, "free": 474, "new": 251, "total": 1729},
    "02 Densa Zuriya Health Post": {"high": 147, "medium": 316, "free": 155, "new": 0, "total": 618},
    "03 Derew Health Post": {"high": 456, "medium": 557, "free": 478, "new": 429, "total": 1920},
    "04 Wejed Health Post": {"high": 246, "medium": 346, "free": 249, "new": 0, "total": 841},
    "06 Gert Health Post": {"high": 237, "medium": 298, "free": 255, "new": 22, "total": 812},
    "07 Lenguat Health Post": {"high": 240, "medium": 328, "free": 244, "new": 0, "total": 812},
    "08 Alegeta Health Post": {"high": 217, "medium": 252, "free": 248, "new": 22, "total": 739},
    "09 Sensa Health Post": {"high": 173, "medium": 272, "free": 179, "new": 0, "total": 624}
}
INSTITUTIONS = list(PLANS.keys())
METRICS = ["High", "Medium", "Free", "New"]

# --- Role-based users (simple local mapping) ---
USERS = {
    "Belay Melaku": {"pwd": "@densa1972", "role": "admin"},
    "viewer": {"pwd": "viewer123", "role": "viewer"}
}

# --- 2. DATABASE CONNECTION ---
def connect_to_gsheets():
    """
    Connect to Google Sheets using service account credentials stored in Streamlit secrets
    under the key 'gcp_service_account'. The secret may be either a dict or a JSON string.
    Returns a gspread client and worksheet objects (records_ws, audit_ws) or (None, None, None) on failure.
    """
    try:
        creds_raw = st.secrets.get("gcp_service_account")
        if not creds_raw:
            st.info("Google Sheets credentials not found in st.secrets['gcp_service_account']. Running in offline/upload mode.")
            return None, None, None

        # Accept dict or JSON string
        if isinstance(creds_raw, str):
            creds_dict = json.loads(creds_raw)
        elif isinstance(creds_raw, dict):
            creds_dict = creds_raw
        else:
            creds_dict = dict(creds_raw)

        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        client = gspread.authorize(creds)
        ss = client.open("CBHI_Data_Database")
        records_ws = ss.worksheet("Records")
        # create or get audit log sheet
        try:
            audit_ws = ss.worksheet("Audit_Log")
        except Exception:
            audit_ws = ss.add_worksheet("Audit_Log", rows=1000, cols=10)
            audit_ws.append_row(["Timestamp", "User", "Role", "Action", "Target", "Details"])
        return client, records_ws, audit_ws
    except Exception:
        st.error("Could not connect to Google Sheets. Running in offline mode.")
        st.exception(traceback.format_exc())
        return None, None, None

# --- 3. UTILITIES ---
def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize column names and types:
    - Recognize 'Report Date' or similar and parse as datetime
    - Ensure metric columns exist and are numeric
    - Ensure 'Health Institution' column exists
    """
    df = df.copy()
    # Strip column whitespace
    df.columns = [c.strip() if isinstance(c, str) else c for c in df.columns]

    # Find date column
    date_cols = [c for c in df.columns if isinstance(c, str) and ("date" in c.lower() or "report" in c.lower())]
    if date_cols:
        date_col = date_cols[0]
        df["Report Date"] = pd.to_datetime(df[date_col], errors="coerce")
    else:
        df["Report Date"] = pd.NaT

    # Health Institution column detection
    inst_cols = [c for c in df.columns if isinstance(c, str) and ("institution" in c.lower() or "health" in c.lower() or "facility" in c.lower())]
    if inst_cols:
        df["Health Institution"] = df[inst_cols[0]].astype(str)
    else:
        if df.shape[1] >= 4:
            df["Health Institution"] = df.iloc[:, 3].astype(str)
        else:
            df["Health Institution"] = ""

    # Metric columns: create/normalize
    for metric in METRICS:
        found = None
        for c in df.columns:
            if isinstance(c, str) and c.strip().lower() == metric.lower():
                found = c
                break
        if found:
            df[metric] = pd.to_numeric(df[found], errors="coerce").fillna(0).astype(int)
        else:
            df[metric] = 0

    return df

def build_plan_df(plans: dict) -> pd.DataFrame:
    rows = []
    for inst, p in plans.items():
        row = {"Health Institution": inst,
               "Plan_High": p.get("high", 0),
               "Plan_Medium": p.get("medium", 0),
               "Plan_Free": p.get("free", 0),
               "Plan_New": p.get("new", 0),
               "Plan_Total": p.get("total", 0)}
        rows.append(row)
    return pd.DataFrame(rows)

def compute_report(plan_df: pd.DataFrame, actual_df: pd.DataFrame, selected_metrics=None):
    """
    Merge plan and actual and compute percentages.
    Returns a wide table with for each institution: Plan_xxx, Actual_xxx, Perc_xxx.
    """
    if selected_metrics is None:
        selected_metrics = METRICS

    # Aggregate actuals by institution
    agg = actual_df.groupby("Health Institution")[METRICS].sum().reset_index()
    # Ensure all institutions appear
    plan_copy = plan_df.copy()
    merged = plan_copy.merge(agg, left_on="Health Institution", right_on="Health Institution", how="left")
    merged.fillna(0, inplace=True)

    # Create columns
    report_rows = []
    for _, r in merged.iterrows():
        inst = r["Health Institution"]
        row = {"Health Institution": inst}
        total_plan = r.get("Plan_Total", 0)
        total_actual = sum(int(r.get(m, 0)) for m in METRICS)
        row["Plan_Total"] = int(total_plan)
        row["Actual_Total"] = int(total_actual)
        row["Perc_Total"] = (total_actual / total_plan * 100) if total_plan else 0
        for metric in selected_metrics:
            p = int(r.get(f"Plan_{metric}", r.get(metric.lower(), 0)))
            a = int(r.get(metric, 0))
            perc = (a / p * 100) if p > 0 else (100.0 if a == 0 else 0.0)
            row[f"Plan_{metric}"] = p
            row[f"Actual_{metric}"] = a
            row[f"Perc_{metric}"] = perc
        report_rows.append(row)

    report_df = pd.DataFrame(report_rows)
    # Add an "Overall" totals row
    totals = {"Health Institution": "ALL INSTITUTIONS",
              "Plan_Total": report_df["Plan_Total"].sum(),
              "Actual_Total": report_df["Actual_Total"].sum(),
              "Perc_Total": (report_df["Actual_Total"].sum() / report_df["Plan_Total"].sum() * 100) if report_df["Plan_Total"].sum() else 0}
    for metric in selected_metrics:
        totals[f"Plan_{metric}"] = report_df[f"Plan_{metric}"].sum()
        totals[f"Actual_{metric}"] = report_df[f"Actual_{metric}"].sum()
        plan_sum = totals[f"Plan_{metric}"]
        actual_sum = totals[f"Actual_{metric}"]
        totals[f"Perc_{metric}"] = (actual_sum / plan_sum * 100) if plan_sum else 0
    report_df = pd.concat([report_df, pd.DataFrame([totals])], ignore_index=True)
    return report_df

def aggregate_time_series(df: pd.DataFrame, metric: str, freq: str):
    """
    Aggregates the provided metric by period.
    freq: 'W' (weekly) or 'M' (monthly)
    Returns a DataFrame with columns ['Period', 'Actual'] where Period is timestamp at period start.
    """
    d = df.dropna(subset=["Report Date"]).copy()
    if d.empty:
        return pd.DataFrame(columns=["Period", "Actual"])
    if freq == "W":
        d["Period"] = d["Report Date"].dt.to_period("W").apply(lambda r: r.start_time)
    else:
        d["Period"] = d["Report Date"].dt.to_period("M").dt.to_timestamp()
    agg = d.groupby("Period")[metric].sum().reset_index().sort_values("Period")
    return agg

def compute_trend_delta(series_df: pd.DataFrame):
    """
    Given a time series DataFrame with Period and Actual, compute delta vs previous period.
    Adds columns Prev and Delta and DeltaPct.
    """
    df = series_df.copy()
    df["Prev"] = df["Actual"].shift(1).fillna(0).astype(int)
    df["Delta"] = df["Actual"] - df["Prev"]
    df["DeltaPct"] = df.apply(lambda r: (r["Delta"] / r["Prev"] * 100) if r["Prev"] else (100.0 if r["Actual"] else 0.0), axis=1)
    return df

def to_excel_with_conditional_formatting(report_df: pd.DataFrame, filename="report.xlsx"):
    """
    Create an Excel bytes buffer with conditional formatting on Perc columns.
    Green >= 70, Yellow 50-70, Red <50
    """
    buffer = BytesIO()
    # Write dataframe to excel
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        report_df.to_excel(writer, index=False, sheet_name="Report")
        writer.save()

    buffer.seek(0)
    if openpyxl is None:
        # If openpyxl not available, return raw buffer
        return buffer

    wb = openpyxl.load_workbook(buffer)
    ws = wb["Report"]

    # Find columns that contain 'Perc' in header
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    perc_cols = [i + 1 for i, h in enumerate(headers) if isinstance(h, str) and ("Perc" in h or "Perc (%)" in h)]
    # Apply formatting to perc cols
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bold = Font(bold=True)

    # header bold
    for cell in ws[1]:
        cell.font = bold

    # iterate rows and apply fill based on value
    for row in range(2, ws.max_row + 1):
        for col in perc_cols:
            cell = ws.cell(row=row, column=col)
            try:
                val = float(cell.value)
            except Exception:
                # Might be string like "70%" - try to parse
                try:
                    vs = str(cell.value).strip().replace("%", "")
                    val = float(vs)
                except Exception:
                    val = None
            if val is not None:
                if val >= 70:
                    cell.fill = green
                elif val >= 50:
                    cell.fill = yellow
                else:
                    cell.fill = red

    # Auto-width columns (simple heuristic)
    for i, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            l = len(str(cell.value))
            if l > max_len:
                max_len = l
        ws.column_dimensions[get_column_letter(i)].width = min(max(50, max_len + 2), 60)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# --- 4. AUDIT LOGGING ---
def append_audit(audit_ws, client, user, role, action, target="", details=""):
    """
    Append audit to Google Sheet if available, else append to local CSV file 'audit_log.csv'.
    """
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = [ts, user, role, action, target, details]
    if audit_ws:
        try:
            audit_ws.append_row(row)
        except Exception:
            # ignore silently but write local fallback
            _append_local_audit(row)
    else:
        _append_local_audit(row)

def _append_local_audit(row):
    fname = "audit_log.csv"
    header = ["Timestamp", "User", "Role", "Action", "Target", "Details"]
    write_header = not os.path.exists(fname)
    try:
        with open(fname, "a", encoding="utf-8") as f:
            if write_header:
                f.write(",".join(header) + "\n")
            # escape commas in fields
            safe = [str(item).replace(",", " ") for item in row]
            f.write(",".join(safe) + "\n")
    except Exception:
        pass

# --- 5. APP LAYOUT & AUTH ---
st.markdown("<h1 style='text-align:center;'>📈 CBHI Performance Tracker</h1>", unsafe_allow_html=True)
st.markdown("Enhanced: weekly/monthly trends, period deltas, conditional Excel formatting, role-based access and audit logging.")

# Connect to Google Sheets (if available)
client, records_ws, audit_ws = connect_to_gsheets()

# Sidebar: authentication
st.sidebar.title("🔐 Sign in")
username = st.sidebar.text_input("Username")
password = st.sidebar.text_input("Password", type="password")

user_role = None
auth_ok = False
if username and password:
    user_info = USERS.get(username)
    if user_info and user_info["pwd"] == password:
        auth_ok = True
        user_role = user_info["role"]
        st.sidebar.success(f"Signed in as {username} ({user_role})")
        append_audit(audit_ws, client, username, user_role, "Login", details="Successful login")
    else:
        st.sidebar.error("Invalid credentials")
        append_audit(audit_ws, client, username or "unknown", "unknown", "Login Failed", details="Invalid credentials")

# Sidebar: data source & upload
st.sidebar.markdown("## Data Source")
data_source = st.sidebar.radio("Choose data source", ("Google Sheets (if connected)", "Upload CSV (offline)"))
uploaded_file = None
if data_source.startswith("Upload"):
    uploaded_file = st.sidebar.file_uploader("Upload CSV of records", type=["csv"], help="CSV should contain Health Institution, High, Medium, Free, New, Report Date")
    if uploaded_file:
        append_audit(audit_ws, client, username or "anonymous", user_role or "unknown", "Upload CSV", details=getattr(uploaded_file, "name", "uploaded"))

def load_data():
    if data_source == "Google Sheets (if connected)":
        if records_ws:
            try:
                records = records_ws.get_all_records()
                df = pd.DataFrame(records)
                df = normalize_df(df)
                return df
            except Exception:
                st.error("Failed to read from Google Sheets.")
                st.exception(traceback.format_exc())
                return pd.DataFrame()
        else:
            st.info("No Google Sheets connection. Please upload CSV in the sidebar.")
            return pd.DataFrame()
    else:
        if uploaded_file is not None:
            try:
                df = pd.read_csv(uploaded_file)
                df = normalize_df(df)
                return df
            except Exception:
                st.error("Failed to parse uploaded CSV.")
                st.exception(traceback.format_exc())
                return pd.DataFrame()
        else:
            return pd.DataFrame()

# Navigation
if auth_ok and user_role == "admin":
    menu = st.sidebar.selectbox("Main Navigation", ["📝 Data Entry", "📊 Dashboard", "🏥 Institution Report", "⚙️ Custom Reports", "🔎 Audit Log"])
else:
    menu = st.sidebar.selectbox("Main Navigation", ["📊 Dashboard", "🏥 Institution Report", "⚙️ Custom Reports"])

# Data Entry (admin)
if menu == "📝 Data Entry" and auth_ok and user_role == "admin":
    st.header("Daily Achievement Entry (Admin)")
    with st.container():
        col1, col2 = st.columns(2)
        reporter = col1.text_input("Reporter Name *")
        phone = col2.text_input("Phone Number *")
        inst = st.selectbox("Health Institution", INSTITUTIONS)
        date_rep = st.date_input("Report Date", datetime.now())

        st.markdown("---")
        st.subheader("1. Membership Achievement Counts")
        r1, r2, r3, r4 = st.columns(4)
        h_val = r1.number_input("High", min_value=0, step=1)
        m_val = r2.number_input("Medium", min_value=0, step=1)
        f_val = r3.number_input("Free", min_value=0, step=1)
        n_val = r4.number_input("New", min_value=0, step=1)

        calc_money = (h_val * 1710) + (m_val * 1260) + (n_val * 1260)

        st.markdown("---")
        st.subheader("2. Financial Status")
        f1, f2 = st.columns(2)
        f1.metric("Calculated Collected (ETB)", f"{calc_money:,.2f}")
        saved = f2.number_input("Actual Amount Saved to Bank (ETB) *", min_value=0.0)

        if st.button("🚀 Submit Final Report"):
            if not reporter or not phone:
                st.error("Please fill in Name and Phone.")
            else:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                new_row = [str(date_rep), reporter, phone, inst, int(h_val), int(m_val), int(f_val), int(n_val), float(calc_money), float(saved), timestamp]
                if records_ws:
                    try:
                        records_ws.append_row(new_row)
                        st.success(f"✅ Success! Report for {inst} has been synced.")
                        st.balloons()
                        append_audit(audit_ws, client, username, user_role, "Submit Report", target=inst, details=f"H:{h_val},M:{m_val},F:{f_val},N:{n_val}")
                    except Exception:
                        st.error("Failed to append to Google Sheets. See details below.")
                        st.exception(traceback.format_exc())
                        _append_local_row = pd.DataFrame([{
                            "Report Date": str(date_rep),
                            "Reporter": reporter,
                            "Phone": phone,
                            "Health Institution": inst,
                            "High": int(h_val),
                            "Medium": int(m_val),
                            "Free": int(f_val),
                            "New": int(n_val),
                            "Calculated Collected (ETB)": float(calc_money),
                            "Saved to Bank (ETB)": float(saved),
                            "Submitted At": timestamp
                        }])
                        csv_buffer = _append_local_row.to_csv(index=False)
                        st.download_button("📥 Download report as CSV", csv_buffer, file_name="cbhi_report_offline.csv", mime="text/csv")
                else:
                    st.warning("No Google Sheets connection detected. Preparing downloadable CSV with the report row.")
                    df_local = pd.DataFrame([{
                        "Report Date": str(date_rep),
                        "Reporter": reporter,
                        "Phone": phone,
                        "Health Institution": inst,
                        "High": int(h_val),
                        "Medium": int(m_val),
                        "Free": int(f_val),
                        "New": int(n_val),
                        "Calculated Collected (ETB)": float(calc_money),
                        "Saved to Bank (ETB)": float(saved),
                        "Submitted At": timestamp
                    }])
                    csv_buffer = df_local.to_csv(index=False)
                    st.download_button("📥 Download report as CSV", csv_buffer, file_name="cbhi_report_offline.csv", mime="text/csv")
                    append_audit(audit_ws, client, username, user_role, "Submit Report (offline)", target=inst, details=f"H:{h_val},M:{m_val},F:{f_val},N:{n_val}")

# Dashboard (all users)
if menu == "📊 Dashboard":
    st.header("Performance Dashboard — Aggregation & Trends")
    df = load_data()

    if df.empty:
        st.info("No data available. Use the sidebar to choose Google Sheets or upload a CSV file.")
    else:
        # Filters & options
        with st.sidebar.expander("Filters & Trend Options", expanded=True):
            insts = st.multiselect("Select Institutions (empty = all)", INSTITUTIONS, default=INSTITUTIONS)
            today = datetime.now().date()
            min_date = df["Report Date"].min().date() if not df["Report Date"].isna().all() else today - timedelta(days=365)
            max_date = df["Report Date"].max().date() if not df["Report Date"].isna().all() else today
            date_range = st.date_input("Date range", value=(min_date, max_date))
            metrics_sel = st.multiselect("Metrics to include", METRICS, default=METRICS)
            agg_freq = st.radio("Time-series aggregation", ("Weekly", "Monthly"))
            show_trend = st.checkbox("Show trend & delta vs previous period", value=True)
            show_charts = st.checkbox("Show charts", value=True)
            highlight_low = st.slider("Highlight institutions below % of plan (Total)", 0, 100, 70)

        # Apply filters
        df_filtered = df.copy()
        if insts:
            df_filtered = df_filtered[df_filtered["Health Institution"].isin(insts)]
        # Date filtering if the Report Date has values
        if not df_filtered["Report Date"].isna().all():
            start_date, end_date = date_range
            mask = (df_filtered["Report Date"].dt.date >= start_date) & (df_filtered["Report Date"].dt.date <= end_date)
            df_filtered = df_filtered[mask]

        plan_df = build_plan_df(PLANS)
        plan_subset = plan_df[plan_df["Health Institution"].isin(insts)] if insts else plan_df
        report_df = compute_report(plan_subset, df_filtered, selected_metrics=metrics_sel)

        # KPI Summary
        st.subheader("Overall KPI Summary")
        kcols = st.columns(4)
        overall = report_df[report_df["Health Institution"] == "ALL INSTITUTIONS"].iloc[0]
        def kpi_widget(col, metric_name):
            plan = overall.get(f"Plan_{metric_name}", 0) if metric_name!="Total" else overall.get("Plan_Total",0)
            actual = overall.get(f"Actual_{metric_name}", 0) if metric_name!="Total" else overall.get("Actual_Total",0)
            perc = overall.get(f"Perc_{metric_name}", 0) if metric_name!="Total" else overall.get("Perc_Total",0)
            col.metric(f"{metric_name} (Act / Plan)", f"{int(actual):,} / {int(plan):,}", f"{perc:.1f}%")
        kpi_widget(kcols[0], "High")
        kpi_widget(kcols[1], "Medium")
        kpi_widget(kcols[2], "Free")
        kpi_widget(kcols[3], "New")

        st.markdown("---")
        st.subheader("Institutional Performance Table")
        display_df = report_df.copy()
        display_df["Status"] = display_df["Perc_Total"].apply(lambda v: "✅ Good" if v >= highlight_low else "⚠️ Low")
        cols_order = ["Health Institution"]
        for metric in metrics_sel:
            cols_order += [f"Plan_{metric}", f"Actual_{metric}", f"Perc_{metric}"]
        cols_order += ["Plan_Total", "Actual_Total", "Perc_Total", "Status"]
        cols_order = [c for c in cols_order if c in display_df.columns]
        st.dataframe(display_df[cols_order].sort_values("Health Institution").reset_index(drop=True), use_container_width=True)

        # Time-series & trends
        if show_charts:
            st.markdown("---")
            st.subheader("Time Series: Actuals by Period")
            freq = "W" if agg_freq == "Weekly" else "M"
            for metric in metrics_sel:
                ts_df = aggregate_time_series(df_filtered, metric, freq)
                ts_trend = compute_trend_delta(ts_df) if show_trend else ts_df
                fig = px.line(ts_trend, x="Period", y="Actual", markers=True, title=f"{metric} — {agg_freq} Actuals")
                # add previous as dotted (Prev is lag, show separate line if desired)
                if show_trend:
                    fig.add_bar(x=ts_trend["Period"], y=ts_trend["Prev"], name="Prev Period", opacity=0.45)
                # annotate delta on points
                if show_trend and not ts_trend.empty:
                    annotations = []
                    for _, r in ts_trend.iterrows():
                        annotations.append(dict(x=r["Period"], y=r["Actual"], text=f"{int(r['Actual']):,} ({r['DeltaPct']:.1f}%)", showarrow=False, yshift=10))
                    fig.update_layout(annotations=annotations)
                fig.update_layout(height=400, xaxis_title="Period", yaxis_title="Count", margin=dict(t=60, b=60))
                st.plotly_chart(fig, use_container_width=True)

        st.markdown("---")
        st.subheader("Export Summary Report")
        csv_bytes = report_df.to_csv(index=False).encode("utf-8")
        st.download_button("📥 Download report as CSV", csv_bytes, file_name="cbhi_report_summary.csv", mime="text/csv")
        try:
            excel_buf = to_excel_with_conditional_formatting(report_df)
            st.download_button("📥 Download report as Excel (with conditional formatting)", excel_buf, file_name="cbhi_report_summary.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception:
            st.info("Excel export requires openpyxl to apply conditional formatting; falling back to plain Excel.")
            # fallback
            with BytesIO() as b:
                with pd.ExcelWriter(b, engine="openpyxl") as writer:
                    report_df.to_excel(writer, index=False, sheet_name="Report")
                b.seek(0)
                st.download_button("📥 Download report as Excel", b, file_name="cbhi_report_summary.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# Institution-only report with numbers on graph, weekly/monthly charts and trend/delta
if menu == "🏥 Institution Report":
    st.header("Institution-Only Report")
    df = load_data()
    if df.empty:
        st.info("No data available. Upload CSV or enable Google Sheets.")
    else:
        inst_choice = st.selectbox("Select Health Institution", ["-- Select --"] + INSTITUTIONS)
        metrics_choice = st.multiselect("Metrics to include", METRICS, default=METRICS)
        min_date = df["Report Date"].min().date() if not df["Report Date"].isna().all() else datetime.now().date() - timedelta(days=365)
        max_date = df["Report Date"].max().date() if not df["Report Date"].isna().all() else datetime.now().date()
        start_date, end_date = st.date_input("Report Date Range", value=(min_date, max_date))
        agg_freq = st.radio("Trend aggregation", ("Weekly", "Monthly"))
        show_numbers_on_chart = st.checkbox("Show numbers on bars", value=True)
        include_percent_labels = st.checkbox("Show percent labels", value=True)
        generate_btn = st.button("Generate")

        if generate_btn:
            if inst_choice == "-- Select --":
                st.error("Please select an institution.")
            else:
                # filter
                df_inst = df[df["Health Institution"] == inst_choice].copy()
                if not df_inst["Report Date"].isna().all():
                    mask = (df_inst["Report Date"].dt.date >= start_date) & (df_inst["Report Date"].dt.date <= end_date)
                    df_inst = df_inst[mask]
                actuals = {m: int(df_inst[m].sum()) for m in METRICS}
                plan_vals = {"High": PLANS.get(inst_choice, {}).get("high", 0),
                             "Medium": PLANS.get(inst_choice, {}).get("medium", 0),
                             "Free": PLANS.get(inst_choice, {}).get("free", 0),
                             "New": PLANS.get(inst_choice, {}).get("new", 0)}
                rows = []
                for m in metrics_choice:
                    p = int(plan_vals.get(m, 0))
                    a = int(actuals.get(m, 0))
                    perc = (a / p * 100) if p > 0 else (100.0 if a == 0 else 0.0)
                    rows.append({"Metric": m, "Plan": p, "Actual": a, "Perc (%)": round(perc, 1)})
                total_plan = sum(r["Plan"] for r in rows)
                total_actual = sum(r["Actual"] for r in rows)
                total_perc = (total_actual / total_plan * 100) if total_plan else 0
                rows.append({"Metric": "TOTAL", "Plan": total_plan, "Actual": total_actual, "Perc (%)": round(total_perc, 1)})
                inst_report_df = pd.DataFrame(rows)

                st.subheader(f"{inst_choice} — Plan vs Actual")
                st.table(inst_report_df.style.format({"Plan": "{:,.0f}", "Actual": "{:,.0f}", "Perc (%)": "{:.1f}%"}))

                # Bar chart numbers
                plot_df = pd.melt(inst_report_df[inst_report_df["Metric"] != "TOTAL"], id_vars=["Metric", "Perc (%)"], value_vars=["Plan", "Actual"], var_name="Series", value_name="Value")
                fig = px.bar(plot_df, x="Metric", y="Value", color="Series", barmode="group", title=f"{inst_choice} — Plan vs Actual (Numbers)")
                if show_numbers_on_chart:
                    fig.update_traces(texttemplate='%{y:,}', textposition='auto')
                if include_percent_labels:
                    annotations = []
                    for idx, row in inst_report_df[inst_report_df["Metric"] != "TOTAL"].iterrows():
                        metric = row["Metric"]
                        perc_val = row["Perc (%)"]
                        annotations.append(dict(x=metric, y=max(row["Plan"], row["Actual"]) * 1.05 + 1, text=f"{perc_val}%", showarrow=False))
                    fig.update_layout(annotations=annotations)
                st.plotly_chart(fig, use_container_width=True)

                # KPI row
                kpcols = st.columns(4)
                for c, metric in zip(kpcols, METRICS):
                    p = int(plan_vals.get(metric, 0))
                    a = int(actuals.get(metric, 0))
                    perc = (a / p * 100) if p > 0 else (100.0 if a == 0 else 0.0)
                    c.metric(f"{metric}", f"{a:,}/{p:,}", f"{perc:.1f}%")

                # Time-series and trend for selected metrics
                freq = "W" if agg_freq == "Weekly" else "M"
                st.markdown("---")
                st.subheader("Time Series & Trend (Institution)")
                for metric in metrics_choice:
                    ts = aggregate_time_series(df_inst, metric, freq)
                    ts_trend = compute_trend_delta(ts) if not ts.empty else ts
                    fig2 = px.line(ts_trend, x="Period", y="Actual", markers=True, title=f"{metric} — {agg_freq} Actuals ({inst_choice})")
                    if not ts_trend.empty:
                        fig2.add_bar(x=ts_trend["Period"], y=ts_trend["Prev"], name="Prev Period", opacity=0.4)
                        annotations = []
                        for _, r in ts_trend.iterrows():
                            annotations.append(dict(x=r["Period"], y=r["Actual"], text=f"{int(r['Actual']):,} ({r['DeltaPct']:.1f}%)", showarrow=False, yshift=10))
                        fig2.update_layout(annotations=annotations)
                    fig2.update_layout(height=400, margin=dict(t=60))
                    st.plotly_chart(fig2, use_container_width=True)

                # Export with formatting
                st.markdown("---")
                csv_bytes = inst_report_df.to_csv(index=False).encode("utf-8")
                st.download_button("📥 Download Institution Report (CSV)", csv_bytes, file_name=f"{inst_choice}_report.csv", mime="text/csv")
                try:
                    excel_buf = to_excel_with_conditional_formatting(inst_report_df, filename=f"{inst_choice}_report.xlsx")
                    st.download_button("📥 Download Institution Report (Excel)", excel_buf, file_name=f"{inst_choice}_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                except Exception:
                    st.info("Excel export requires openpyxl to apply conditional formatting.")

                append_audit(audit_ws, client, username or "anonymous", user_role or "unknown", "Generate Institution Report", target=inst_choice, details=f"Metrics:{','.join(metrics_choice)}; Range:{start_date} to {end_date}")

# Custom reports (all users)
if menu == "⚙️ Custom Reports":
    st.header("Custom Report Generator")
    df = load_data()
    if df.empty:
        st.info("No data available.")
    else:
        with st.form("custom_report"):
            colA, colB = st.columns(2)
            sel_insts = colA.multiselect("Institutions (leave empty = all)", INSTITUTIONS, default=INSTITUTIONS)
            sel_metrics = colB.multiselect("Metrics to include", METRICS, default=METRICS)
            min_date = df["Report Date"].min().date() if not df["Report Date"].isna().all() else datetime.now().date() - timedelta(days=365)
            max_date = df["Report Date"].max().date() if not df["Report Date"].isna().all() else datetime.now().date()
            drange = st.date_input("Report Date Range", value=(min_date, max_date))
            output_format = st.selectbox("Export format", ["CSV", "Excel (with conditional formatting)"])
            include_charts = st.checkbox("Include charts in the app view", value=True)
            submitted = st.form_submit_button("Generate")

        if submitted:
            df_filtered = df.copy()
            if sel_insts:
                df_filtered = df_filtered[df_filtered["Health Institution"].isin(sel_insts)]
            if not df_filtered["Report Date"].isna().all():
                start_date, end_date = drange
                mask = (df_filtered["Report Date"].dt.date >= start_date) & (df_filtered["Report Date"].dt.date <= end_date)
                df_filtered = df_filtered[mask]
            plan_df = build_plan_df(PLANS)
            plan_df = plan_df[plan_df["Health Institution"].isin(sel_insts)] if sel_insts else plan_df
            report_df = compute_report(plan_df, df_filtered, selected_metrics=sel_metrics)

            st.success("Report generated")
            st.dataframe(report_df, use_container_width=True)

            if include_charts:
                for metric in sel_metrics:
                    fig = px.bar(
                        report_df[report_df["Health Institution"] != "ALL INSTITUTIONS"],
                        x="Health Institution",
                        y=[f"Plan_{metric}", f"Actual_{metric}"],
                        barmode="group",
                        title=f"{metric} — Plan vs Actual by Institution",
                        labels={"value": metric, "variable": "Series"}
                    )
                    fig.update_traces(texttemplate='%{y:,}', textposition='auto')
                    fig.update_layout(height=400, xaxis_tickangle=-45, margin=dict(t=50, b=120))
                    st.plotly_chart(fig, use_container_width=True)

            # Export
            append_audit(audit_ws, client, username or "anonymous", user_role or "unknown", "Generate Custom Report", details=f"Inst:{','.join(sel_insts) or 'ALL'}; Metrics:{','.join(sel_metrics)}")
            if output_format.startswith("Excel") and openpyxl:
                excel_buf = to_excel_with_conditional_formatting(report_df)
                st.download_button("📥 Download Excel", excel_buf, file_name="cbhi_custom_report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                csv_bytes = report_df.to_csv(index=False).encode("utf-8")
                st.download_button("📥 Download CSV", csv_bytes, file_name="cbhi_custom_report.csv", mime="text/csv")

# Audit Log view (admin only)
if menu == "🔎 Audit Log" and auth_ok and user_role == "admin":
    st.header("Audit Log")
    st.markdown("Shows recent actions (login, report generation, uploads, data entry). Admin-only view.")
    # try to fetch audit sheet
    if audit_ws:
        try:
            logs = audit_ws.get_all_records()
            log_df = pd.DataFrame(logs)
            if log_df.empty:
                st.info("No audit records yet.")
            else:
                st.dataframe(log_df.sort_values("Timestamp", ascending=False).reset_index(drop=True), use_container_width=True)
        except Exception:
            st.error("Failed to read audit sheet.")
            st.exception(traceback.format_exc())
    else:
        # try local file
        if os.path.exists("audit_log.csv"):
            log_df = pd.read_csv("audit_log.csv")
            st.dataframe(log_df.sort_values("Timestamp", ascending=False).reset_index(drop=True), use_container_width=True)
        else:
            st.info("No audit log available.")

# Footer/helper
st.markdown("---")
st.markdown("Need more customizations? I can add PDF export, deeper role management (LDAP/SSO), or push automated weekly summary emails.")